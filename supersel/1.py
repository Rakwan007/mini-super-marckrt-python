import tkinter as tk
from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook
#############################

wb = Workbook()
ws = wb.active
ws.title= 'customer'
ws["A1"]= 'Full Name'
ws["B1"]= 'Number Phone'
ws["C1"]= 'Country'
ws["D1"]= 'Total'
ws["E1"]= 'Date buy'
wb.save('customer.xlsx')

#############################
timenow = datetime.datetime.now()
date = timenow.strftime("%Y-%m-%d")
#############################

root = tk.Tk()
root.geometry("950x552")
root.iconbitmap('img/icon.ico')
root.title('Building Tools [متجر ادوات البناء] ')

menu={
    0:['منشار',20],
    1:['عربة',40],
    2:['معول',30],
    3:['مجرفة',20],
    4:['مطرقة',30],
    5:['دلو',20],
    6:['قبعة امان',40],
    7:['مشرط',30],
    8:['قطاعة',10],
    9:['كماشة',20],
    10:['مفك براغي',30],
    11:['دريل',140],
    }
def test():
    for Parent in trv.get_children():
        print(trv.item(Parent)["text"])
        for child in trv.get_children(Parent):
            data = trv.item(child)["text"]
            print(data)
def my_reset():
    for item in trv.get_children():
        trv.delete(item)
    En_total.delete('0', END)
    En_date.delete('0', END)
    En_name.delete('0', END)
    En_phone.delete('0', END)
    En_address.delete('0', END)
    
def my_bill():
    
    global En_name
    global En_phone
    global En_address
    global En_total
    global En_date
    lb_image.place(x=950,y=438, width=250 , height=110)
    
    root.geometry("1205x552")
    F4 = Frame(root,bg='#5F7161',width=250,height=434,bd=2,relief=GROOVE)
    F4.place(x=950,y=1)

    L_name = Label(F4,text='اسم المشتري',bg='#5F7161',fg='white')
    L_name.place(x=168,y=10)
    En_name = Entry(F4,width=24,font=('Tajawal',12),justify=CENTER)
    En_name.place(x=15,y=40)

    L_phone = Label(F4,text='رقم المشتري ',bg='#5F7161',fg='white')
    L_phone.place(x=170,y=70)
    En_phone = Entry(F4,width=24,font=('Tajawal',12),justify=CENTER)
    En_phone.place(x=15,y=100)

    L_address = Label(F4,text='عنوان المشتري ',bg='#5F7161',fg='white')
    L_address.place(x=160,y=130)
    En_address = Entry(F4,width=24,font=('Tajawal',12),justify=CENTER)
    En_address.place(x=15,y=160)

    L_total = Label(F4,text='الحساب الكلي',bg='#5F7161',fg='white')
    L_total.place(x=165,y=190)
    En_total = Entry(F4,width=24,font=('Tajawal',12),justify=CENTER)
    En_total.place(x=15,y=210)

    L_date = Label(F4,text='تاريخ الشراء',bg='#5F7161',fg='white')
    L_date.place(x=175,y=240)
    En_date = Entry(F4,width=24,font=('Tajawal',12),justify=CENTER)
    En_date.place(x=15,y=270)

    add_button = Button(F4,text='حفظ فاتورة',width=31,cursor='hand2',bg='#EDDBC0',command=buy)
    add_button.place(x=12,y=310)

    add_button = Button(F4,text='افراغ الحقول',width=31,cursor='hand2',bg='#EDDBC0')
    add_button.place(x=12,y=340)

    add_button = Button(F4,text='بحث عن مشتري',width=31,cursor='hand2',bg='#EDDBC0')
    add_button.place(x=12,y=370)

    add_button = Button(F4,text='حذف فاتورة',width=31,cursor='hand2',bg='#EDDBC0')
    add_button.place(x=12,y=400)

    total=0
    for item in trv.get_children():
        trv.delete(item)
    for i in range(len(sb)):
        if(int(sb[i].get())>0):
            price=int(sb[i].get())*menu[i][1]
            total=total+price
            my_str1=(str(menu[i][1]), str(sb[i].get()), str(price))
            trv.insert("",'end',iid=i,text=menu[i][0],values=my_str1)
    final=total
    En_total.insert('1',str(final) + ' $')
    En_date.insert('1',str(date))

def buy():
    name = En_name.get()
    phone = En_phone.get()
    address = En_address.get()
    total = En_total.get()
    datebuy = En_date.get()

    excel = openpyxl.load_workbook('customer.xlsx')
    file = excel.active
    file.cell(column=1,row=file.max_row+1,value=name)
    file.cell(column=2,row=file.max_row,value=phone)
    file.cell(column=3,row=file.max_row,value=address)
    file.cell(column=4,row=file.max_row,value=total)
    file.cell(column=5,row=file.max_row,value=datebuy)
    excel.save('customer.xlsx')
    
#===== Frame[1] ======
F1 = Frame(root,bg='silver',width=600,height=550)
F1.place(x=1,y=1)

#===== Images =====
img_menu1=PhotoImage(file="img/1.png")
img_menu2=PhotoImage(file="img/2.png")
img_menu3=PhotoImage(file="img/3.png")
img_menu4=PhotoImage(file="img/4.png")

img_menu5=PhotoImage(file="img/5.png")
img_menu6=PhotoImage(file="img/6.png")
img_menu7=PhotoImage(file="img/7.png")
img_menu8=PhotoImage(file="img/8.png")

img_menu9=PhotoImage(file="img/9.png")
img_menu10=PhotoImage(file="img/10.png")
img_menu11=PhotoImage(file="img/11.png")
img_menu12=PhotoImage(file="img/12.png")


title = Label(F1,text='مشروع بيع معدات بناء',font=('Tajawal 13'),fg='white',bg='#5F7161',width=70)
title.place(x=0,y=0)
menu1=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu1,text='منشار',compound=TOP)
menu1.place(x=30,y=45)
menu2=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu2,text='عربة',compound=TOP)
menu2.place(x=170,y=45)
menu3=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu3,text='معول',compound=TOP)
menu3.place(x=310,y=45)
menu4=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu4,text='مجرفة',compound=TOP)
menu4.place(x=450,y=45)

menu5=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu5,text='مطرقة',compound=TOP)
menu5.place(x=30,y=180)
menu6=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu6,text='دلو',compound=TOP)
menu6.place(x=170,y=180)
menu7=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu7,text='قبعة امان',compound=TOP)
menu7.place(x=310,y=180)
menu8=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu8,text='مشرط',compound=TOP)
menu8.place(x=450,y=180)

menu9=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu9,text='قطاعة',compound=TOP)
menu9.place(x=30,y=320)
menu10=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu10,text='بنسة',compound=TOP)
menu10.place(x=170,y=320)
menu11=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu11,text='مفك براغي',compound=TOP)
menu11.place(x=310,y=320)
menu12=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu12,text='دريل',compound=TOP)
menu12.place(x=450,y=320)
#====== Variable =======
sb=[]
font1=('Times',12,'normal')
font2=('Times',32,'bold')
pdx,pdy=40,5

sv1=IntVar()
sb1 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv1)
sb1.place(x=30,y=140)
sb.append(sb1)    
sv2=IntVar()
sb2 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv2)
sb2.place(x=170,y=140)
sb.append(sb2)    
sv3=IntVar()
sb3 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv3)
sb3.place(x=310,y=140)
sb.append(sb3)    
sv4=IntVar()
sb4 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv4)
sb4.place(x=450,y=140)
sb.append(sb4)

sv5=IntVar()
sb5 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv5)
sb5.place(x=30,y=275)
sb.append(sb5)
sv6=IntVar()
sb6 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv6)
sb6.place(x=170,y=275)
sb.append(sb6)
sv7=IntVar()
sb7 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv7)
sb7.place(x=310,y=275)
sb.append(sb7)
sv8=IntVar()
sb8 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv8)
sb8.place(x=450,y=275)
sb.append(sb8)

sv9=IntVar()
sb9 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv9)
sb9.place(x=30,y=415)
sb.append(sb9)    
sv10=IntVar()
sb10 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv10)
sb10.place(x=170,y=415)
sb.append(sb10)    
sv11=IntVar()
sb11 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv11)
sb11.place(x=310,y=415)
sb.append(sb11)    
sv12=IntVar()
sb12 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv12)
sb12.place(x=450,y=415)
sb.append(sb12)


#======== Frame[2]==========
F2 = Frame(root,bg='gray',width=343,height=550)
F2.place(x=604,y=1)

trv = ttk.Treeview(F2, selectmode ='browse')
trv.place(x=1,y=1,width=340,height=550)
trv["columns"] = ("1", "2","3")
trv.column("#0", width = 80, anchor ='c')
trv.column("1", width = 50, anchor ='c')
trv.column("2", width =50 , anchor ='c')
trv.column("3", width = 60, anchor ='c')
trv.heading("#0", text ="المواد",anchor='c')
trv.heading("1", text ="السعر",anchor='c')
trv.heading("2", text ="العدد",anchor='c')
trv.heading("3", text ="الحساب الكلي",anchor='c')


#======== Frame[3]==========



#====== Button Rent ======
b1=Button(F1,text='شراء المواد',fg='white',font=('Tajawal 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command=my_bill)
b1.place(x=30,y=500)
b2=Button(F1,text='فاتورة جديدة',fg='white',font=('Tajawal 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command=my_reset)
b2.place(x=160,y=500)
b3=Button(F1,text=' استئجار مواد',fg='white',font=('Tajawal 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1)
b3.place(x=290,y=500)
b4=Button(F1,text=' اغلاق البرنامج',fg='white',font=('Tajawal 12'),width=15,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,command=test)
b4.place(x=420,y=500)


im_logo = PhotoImage(file="img/logo.png")
lb_image= Label(root,image=im_logo)


root.mainloop()